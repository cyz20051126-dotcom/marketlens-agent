from __future__ import annotations

import csv
import hashlib
import json
import re
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import date
from pathlib import Path
from urllib.parse import urlparse

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
FIRECRAWL_DIR = ROOT / ".firecrawl" / "jobs"
OUT_DIR = ROOT / "deliverables" / "jobs"

RAW_CSV = OUT_DIR / "internship_candidates_raw.csv"
TOP_CSV = OUT_DIR / "internship_shortlist_100.csv"
TOP_XLSX = OUT_DIR / "internship_shortlist_100.xlsx"
TOP_MD = OUT_DIR / "internship_shortlist_100.md"

TODAY = date(2026, 6, 19).isoformat()


POSITIVE_TERMS = {
    "AI/Agent/运营": [
        "ai运营",
        "ai 运营",
        "ai产品运营",
        "ai 产品运营",
        "aigc",
        "agent运营",
        "agent 运营",
        "prompt",
        "工作流",
        "ai应用",
        "ai 应用",
        "ai产品",
        "ai 产品",
    ],
    "商业/数据分析": [
        "商业分析",
        "business analyst",
        "数据分析",
        "数据运营",
        "经营分析",
        "策略分析",
        "策略运营",
        "市场分析",
        "市场研究",
        "用户研究",
    ],
    "品牌/增长/内容": [
        "品牌运营",
        "用户运营",
        "内容运营",
        "产品运营",
        "增长",
        "新媒体",
        "小红书",
        "抖音",
        "营销",
        "市场部",
    ],
    "金融/研究": [
        "投研",
        "行业研究",
        "研究助理",
        "金融",
        "基金",
        "证券",
        "行研",
        "财务分析",
        "investment",
        "research intern",
    ],
}

LOCATION_TERMS = ["成都", "远程", "remote", "线上", "居家", "不限地点", "地点不限"]
PAY_TERMS = [
    "元/天",
    "元/日",
    "k",
    "薪",
    "工资待遇",
    "薪资",
    "有薪",
    "日薪",
    "月薪",
    "津贴",
]
AVAILABILITY_TERMS = ["暑期", "7月", "七月", "每周", "3天", "4天", "5天", "8周", "3个月", "实习时间"]

NEGATIVE_TERMS = [
    "算法",
    "后端",
    "前端",
    "java",
    "c++",
    "硬件",
    "固件",
    "sdk",
    "架构开发",
    "系统开发",
    "推荐算法",
    "测试开发",
    "深度学习",
    "博士",
    "硕士优先",
    "新闻",
    "教程",
    "使用工作流",
    "博客",
    "双选会",
]

JUNK_TITLE_TERMS = [
    "![]",
    "base64",
    "登录",
    "注册",
    "用户协议",
    "隐私政策",
    "实习证明",
    "提供实习",
    "实习津贴",
    "职位校招",
    "上传简历",
    "搜索工作",
    "标题：姓名",
    "每周能够",
    "尽快入职",
    "地铁周边",
    "周末双休",
    "刷新",
    "招聘信息",
    "薪资情况",
    "job post",
    "计算机维护",
    "仓库操作员",
    "高级理财经理",
    "高级计划员",
    "chemical analysis",
    "scientist",
    "solutions architect",
    "supervisor",
    "业务发展主管",
    "软件测试",
    "测试工程师",
    "测试开发",
    "全栈开发",
    "ai全栈",
    "数据研发工程师",
    "地产实习生",
    "猎头",
    "生产实习生",
    "跳到内容",
    "该职位将在",
    "招聘职位",
    "招聘专场",
    "运维",
    "engineer",
    "博士",
    "人事实习",
    "人力资源",
    "销售实习",
    "intern_ts",
    "early career",
    "实习生 intern",
    "training",
    "权威运营",
    "开发岗",
    "开发实习",
    "开发（",
    "工程师",
    "硕士",
    "!",
]

ROLE_TERMS = [
    "运营",
    "分析",
    "研究",
    "投研",
    "行研",
    "品牌",
    "产品",
    "市场",
    "增长",
    "用户",
    "内容",
    "商业",
    "数据",
    "ai",
    "aigc",
    "agent",
    "intern",
    "实习生",
    "助理",
]

CORE_TITLE_TERMS = [
    "aigc",
    "agent",
    "运营",
    "商业分析",
    "数据分析",
    "经营分析",
    "策略",
    "品牌",
    "内容",
    "用户",
    "市场",
    "marketing",
    "finance",
    "business",
    "投研",
    "行业研究",
    "研究助理",
    "财务分析",
    "基金",
    "金融",
    "项目管理",
    "产品",
    "channel management",
]

SOURCE_QUALITY = {
    "jobs.bytedance.com": 12,
    "join.qq.com": 12,
    "careers.tencent.com": 12,
    "talent.baidu.com": 12,
    "jobs.bilibili.com": 11,
    "zhaopin.meituan.com": 11,
    "zhaopin.kuaishou.cn": 11,
    "campus.hr.xiaomi.com": 11,
    "www.pwccn.com": 11,
    "campus.51job.com": 9,
    "www.shixiseng.com": 9,
    "wap.shixiseng.com": 9,
    "cn.indeed.com": 8,
    "www.zhipin.com": 7,
    "m.zhipin.com": 7,
    "www.liepin.com": 7,
    "m.liepin.com": 7,
    "www.wondercv.com": 7,
    "career.cuhk.edu.cn": 6,
    "career.nankai.edu.cn": 6,
    "career.uibe.edu.cn": 6,
}


@dataclass
class Candidate:
    title: str
    company: str
    url: str
    source_file: str
    source_domain: str
    snippet: str
    extracted_from: str
    source_type: str
    score: int = 0
    categories: list[str] = field(default_factory=list)
    fit_reason: str = ""
    risk: str = ""
    salary_hint: str = ""
    location_hint: str = ""
    time_hint: str = ""

    def as_row(self) -> dict[str, str | int]:
        return {
            "rank_score": self.score,
            "category": " / ".join(self.categories),
            "title": self.title,
            "company_or_source": self.company,
            "location_remote": self.location_hint,
            "salary": self.salary_hint,
            "time_requirement": self.time_hint,
            "fit_reason": self.fit_reason,
            "risk_or_note": self.risk,
            "source_domain": self.source_domain,
            "source_type": self.source_type,
            "source_url": self.url,
            "extracted_from": self.extracted_from,
            "snippet": clean_space(self.snippet)[:520],
            "collected_at": TODAY,
        }


def clean_space(text: str) -> str:
    return re.sub(r"\s+", " ", (text or "").replace("&nbsp;", " ")).strip()


def domain(url: str) -> str:
    try:
        return urlparse(url).netloc.lower()
    except Exception:
        return ""


def normalize_title(title: str) -> str:
    text = title.lower()
    text = re.sub(r"[\s_\-—|【】\[\]（）()]", "", text)
    text = re.sub(r"202[0-9].*", "", text)
    return text[:80]


def title_from_search(title: str) -> str:
    title = clean_title(title)
    title = re.sub(r"^\d+\s*个", "", title)
    title = re.sub(r"[，,]\s*2026年.*$", "", title)
    title = title.replace("招聘信息】", "招聘信息")
    for sep in [" - ", " | ", "_"]:
        if sep in title and len(title.split(sep)[0]) >= 4:
            left = title.split(sep)[0].strip()
            if not any(x in left for x in ["招聘", "职位", "就业机会"]):
                return left
    return title.strip()


def clean_title(title: str) -> str:
    title = clean_space(title)
    title = re.sub(r"^#+\s*", "", title)
    title = re.sub(r"^.*正在招聘", "", title)
    title = title.strip("[]\\| ")
    title = title.replace("\\", "")
    title = re.sub(r"\s*\|\s*.*$", "", title) if title.startswith("[") else title
    title = title.strip("[]\\| ")
    title = re.sub(r"\s+-\s+加入字节跳动$", "", title)
    title = re.sub(r"\s+-\s*实习僧$", "", title)
    title = re.sub(r"\s+-\s*领英$", "", title)
    title = re.sub(r"招聘$", "", title)
    return title


def is_bad_title(title: str) -> bool:
    text = clean_space(title).lower()
    if not text or len(text) < 3:
        return True
    if any(term.lower() in text for term in JUNK_TITLE_TERMS):
        return True
    if re.fullmatch(r"\d{2,5}\s*-\s*\d{2,5}\s*/?\s*(元|天|日|k|K).+", text):
        return True
    if re.fullmatch(r"\d+\s*-\s*\d+\s*/?\s*天.+", text):
        return True
    if text.endswith("职位") or "职位 - " in text:
        return True
    if len(text) > 90:
        return True
    return False


def looks_like_role(title: str) -> bool:
    text = clean_space(title).lower()
    if is_bad_title(text):
        return False
    return ("实习" in text or "intern" in text) and any(term.lower() in text for term in ROLE_TERMS)


def title_has_core_fit(title: str) -> bool:
    text = clean_space(title).lower()
    if re.search(r"(^|[^a-z])ai([^a-z]|$)", text):
        return True
    return any(term.lower() in text for term in CORE_TITLE_TERMS)


def company_from_text(title: str, text: str, dom: str) -> str:
    source_map = {
        "jobs.bytedance.com": "字节跳动",
        "join.qq.com": "腾讯",
        "careers.tencent.com": "腾讯",
        "talent.baidu.com": "百度",
        "jobs.bilibili.com": "哔哩哔哩",
        "zhaopin.meituan.com": "美团",
        "zhaopin.kuaishou.cn": "快手",
        "campus.hr.xiaomi.com": "小米",
        "www.pwccn.com": "普华永道",
    }
    if dom in source_map:
        return source_map[dom]

    patterns = [
        r"(.{2,30})正在招聘",
        r"公司[:：]\s*([^，。；;\n]{2,30})",
        r"企业[:：]\s*([^，。；;\n]{2,30})",
    ]
    sample = f"{title} {text}"
    for pat in patterns:
        m = re.search(pat, sample)
        if m:
            return clean_space(m.group(1))

    lines = [clean_space(x) for x in text.splitlines() if clean_space(x)]
    bad = ("职位", "岗位", "职责", "要求", "薪资", "工作", "实习", "刷新", "Base64", "![]", "微信扫码")
    for line in lines[:8]:
        if 3 <= len(line) <= 36 and not any(b in line for b in bad):
            return line

    return source_map.get(dom, dom or "未知")


def extract_hints(text: str) -> tuple[str, str, str]:
    salary = ""
    salary_patterns = [
        r"\d{2,5}\s*-\s*\d{2,5}\s*元\s*/\s*[天日]",
        r"\d{2,5}\s*元\s*/\s*[天日]",
        r"\d+\s*-\s*\d+\s*k(?:·\d+薪)?",
        r"\d+\s*k(?:·\d+薪)?",
        r"工资待遇[,，:：]?\s*[^;；。|\n]{1,30}",
        r"薪资待遇[,，:：]?\s*[^;；。|\n]{1,30}",
        r"日薪\s*\d{2,5}[^;；。|\n]{0,12}",
    ]
    for pat in salary_patterns:
        m = re.search(pat, text, flags=re.I)
        if m:
            salary = clean_space(m.group(0))
            break

    loc_hits = []
    for term in LOCATION_TERMS + ["北京", "上海", "深圳", "杭州", "广州", "南京", "苏州"]:
        if term.lower() in text.lower() and term not in loc_hits:
            loc_hits.append(term)
    location = " / ".join(loc_hits[:4])

    time_hits = []
    time_patterns = [
        r"每周[^，。；;\n]{0,20}[345]天",
        r"实习[^，。；;\n]{0,20}[238]周",
        r"实习[^，。；;\n]{0,20}[346]个月",
        r"2026年7-8月",
        r"7月[^，。；;\n]{0,16}",
        r"暑期[^，。；;\n]{0,20}",
    ]
    for pat in time_patterns:
        for m in re.finditer(pat, text):
            hit = clean_space(m.group(0))
            if hit and hit not in time_hits:
                time_hits.append(hit)
    time_hint = " / ".join(time_hits[:3])
    return salary, location, time_hint


def score_candidate(c: Candidate) -> Candidate:
    c.title = clean_title(c.title)
    text = clean_space(f"{c.title} {c.company} {c.snippet}").lower()
    score = 0
    categories = []
    reason_bits = []
    risk_bits = []

    if "实习" in text or "intern" in text:
        score += 18
    else:
        score -= 18
        risk_bits.append("未明确为实习岗位")

    for cat, terms in POSITIVE_TERMS.items():
        hits = [term for term in terms if term.lower() in text]
        if hits:
            categories.append(cat)
            score += min(22, 8 + 4 * len(hits))
            reason_bits.append(f"匹配{cat}关键词：{', '.join(hits[:4])}")

    loc_hits = [term for term in LOCATION_TERMS if term.lower() in text]
    if loc_hits:
        score += 12
        reason_bits.append("地点/形式友好：" + "、".join(loc_hits[:3]))
    elif any(city in text for city in ["北京", "上海", "深圳", "杭州", "广州"]):
        score += 2
        risk_bits.append("异地岗位，需确认是否可远程或暑期线下")
    else:
        risk_bits.append("地点灵活性未知")

    if any(term.lower() in text for term in PAY_TERMS):
        score += 8
        reason_bits.append("页面出现薪资/补贴信息")
    else:
        risk_bits.append("薪资未在摘要中明确")

    if any(term.lower() in text for term in AVAILABILITY_TERMS):
        score += 6
        reason_bits.append("出现暑期/到岗/每周出勤信息")

    if "本科" in text or "在校" in text or "经验不限" in text:
        score += 7
        reason_bits.append("学历或经验门槛相对友好")

    if "2028" in text:
        score += 7
        reason_bits.append("明确覆盖2028届")
    elif "2027" in text:
        score -= 5
        risk_bits.append("可能偏2027届，需要确认是否接受2028届")
    elif "2026" in text:
        score -= 8
        risk_bits.append("可能偏2026届/已毕业，需核验投递资格")

    old_date = re.search(r"202[0-4][-/年]", text) or re.search(r"2025[-/]0[1-5]", text)
    if old_date:
        score -= 35
        risk_bits.append("页面疑似旧岗位/旧刷新时间")

    negative_hits = [term for term in NEGATIVE_TERMS if term.lower() in text]
    if negative_hits:
        penalty = min(28, 6 * len(negative_hits))
        score -= penalty
        risk_bits.append("可能不匹配：" + "、".join(negative_hits[:5]))

    score += SOURCE_QUALITY.get(c.source_domain, 3)

    if len(clean_space(c.snippet)) < 30:
        score -= 8
        risk_bits.append("信息过少，需点开确认")

    c.salary_hint, c.location_hint, c.time_hint = extract_hints(f"{c.title}\n{c.company}\n{c.snippet}")
    c.score = score
    c.categories = categories or ["待核验"]
    c.fit_reason = "；".join(reason_bits[:4]) or "关键词较弱，保留为补充候选"
    c.risk = "；".join(dict.fromkeys(risk_bits)) or "需核验最新开放状态"
    return c


def extract_indeed_jobs(markdown: str, source_url: str, source_file: str) -> list[Candidate]:
    candidates = []
    pattern = re.compile(r"### \[([^\]]{2,120})\]\((https://cn\.indeed\.com/[^\)]+)\)(.*?)(?=\n\| ### \[|\n### \[|\Z)", re.S)
    for m in pattern.finditer(markdown or ""):
        title = clean_space(re.sub(r"<br>", " ", m.group(1)))
        body = clean_space(re.sub(r"<[^>]+>", " ", m.group(3)))
        if not title or len(title) > 80:
            continue
        dom = domain(m.group(2))
        company = company_from_text(title, body, dom)
        candidates.append(
            Candidate(
                title=title,
                company=company,
                url=m.group(2),
                source_file=source_file,
                source_domain=dom,
                snippet=body,
                extracted_from=source_url,
                source_type="indeed-listing",
            )
        )
    return candidates


def extract_markdown_list_jobs(markdown: str, source_url: str, source_file: str, dom: str) -> list[Candidate]:
    candidates = []
    # Avoid extracting page chrome from a single job detail page; the search-result
    # candidate already represents that page.
    if "/intern/inn_" in source_url or "position/detail" in source_url or "rc/clk" in source_url:
        return candidates
    # Generic markdown links that look like internship job titles.
    for title, url in re.findall(r"\[([^\]]{4,80}(?:实习|Intern|intern)[^\]]{0,80})\]\((https?://[^\)]+)\)", markdown or ""):
        title = clean_space(title)
        if not looks_like_role(title):
            continue
        if any(bad in url for bad in ["/rule", "/login", "/resume", "privacy"]):
            continue
        if any(bad in url.lower() for bad in ["image", ".png", ".jpg", ".jpeg", ".gif", "lietou-static"]):
            continue
        if url.rstrip("/") == "https://www.shixiseng.com":
            continue
        around = ""
        pos = markdown.find(title)
        if pos >= 0:
            around = clean_space(markdown[max(0, pos - 260): pos + 520])
        u = url if url.startswith("http") else source_url
        candidates.append(
            Candidate(
                title=title_from_search(title),
                company=company_from_text(title, around, domain(u) or dom),
                url=u,
                source_file=source_file,
                source_domain=domain(u) or dom,
                snippet=around,
                extracted_from=source_url,
                source_type="markdown-link",
            )
        )

    # Plain list entries from dynamic campus pages.
    for line in (markdown or "").splitlines():
        s = clean_space(line.strip("-·* "))
        if not looks_like_role(s):
            continue
        if 4 <= len(s) <= 60 and ("实习" in s or "Intern" in s or "intern" in s):
            candidates.append(
                Candidate(
                    title=title_from_search(s),
                    company=company_from_text(s, markdown[:600], dom),
                    url=source_url,
                    source_file=source_file,
                    source_domain=dom,
                    snippet=clean_space(markdown[:900]),
                    extracted_from=source_url,
                    source_type="page-list-entry",
                )
            )
    return candidates


def load_candidates() -> list[Candidate]:
    candidates: list[Candidate] = []
    for path in sorted(FIRECRAWL_DIR.glob("search-*.json")):
        data = json.loads(path.read_text(encoding="utf-8"))
        for item in (data.get("data") or {}).get("web", []):
            url = item.get("url") or ""
            if not url:
                continue
            dom = domain(url)
            title = title_from_search(item.get("title") or "")
            markdown = item.get("markdown") or ""
            desc = item.get("description") or ""
            snippet = desc or markdown[:1000]
            candidates.append(
                Candidate(
                    title=title,
                    company=company_from_text(title, markdown or desc, dom),
                    url=url,
                    source_file=path.name,
                    source_domain=dom,
                    snippet=snippet,
                    extracted_from=url,
                    source_type="search-result",
                )
            )
            if "cn.indeed.com" in dom and markdown:
                candidates.extend(extract_indeed_jobs(markdown, url, path.name))
            if markdown:
                candidates.extend(extract_markdown_list_jobs(markdown, url, path.name, dom))
    return [score_candidate(c) for c in candidates]


def dedupe(candidates: list[Candidate]) -> list[Candidate]:
    best: dict[str, Candidate] = {}
    for c in candidates:
        c.title = clean_title(c.title)
        title_key = normalize_title(c.title)
        company_key = normalize_title(c.company)
        parsed = urlparse(c.url)
        url_key = f"{parsed.netloc.lower()}{parsed.path}".rstrip("/")
        key = hashlib.md5(f"{title_key}|{company_key}|{url_key}".encode("utf-8")).hexdigest()
        if any(marker in c.url for marker in ["position/detail", "job_detail", "/intern/inn_", "/rc/clk"]):
            key = hashlib.md5(url_key.encode("utf-8")).hexdigest()
        # Also collapse obvious repeated search-result list pages by title/company.
        if c.source_type in {"page-list-entry", "markdown-link", "indeed-listing"}:
            key = hashlib.md5(title_key.encode("utf-8")).hexdigest()
        if key not in best or c.score > best[key].score or len(c.snippet) > len(best[key].snippet):
            best[key] = c
    return sorted(best.values(), key=lambda x: (x.score, x.source_type != "search-result"), reverse=True)


def filter_shortlist(candidates: list[Candidate]) -> list[Candidate]:
    blocked_domains = {
        "www.feishu.cn",
        "aws.amazon.com",
        "www.news.cn",
        "zhuanlan.zhihu.com",
        "image0.lietou-static.com",
        "image1.lietou-static.com",
    }
    shortlist = []
    for c in candidates:
        c.title = clean_title(c.title)
        text = f"{c.title} {c.snippet}".lower()
        if is_bad_title(c.title):
            continue
        if not looks_like_role(c.title):
            continue
        if not title_has_core_fit(c.title):
            continue
        if c.source_type == "page-list-entry" and c.score < 62:
            continue
        if c.source_type == "search-result" and c.source_domain in {
            "www.liepin.com",
            "m.liepin.com",
            "www.zhipin.com",
            "m.zhipin.com",
            "cn.indeed.com",
        } and ("招聘信息" in c.title or "职位" in c.title or "薪资" in c.title):
            continue
        if c.source_type == "search-result" and c.source_domain == "cn.indeed.com" and "/career/" in c.url:
            continue
        if c.url.rstrip("/") in {"https://www.shixiseng.com", "https://wap.shixiseng.com"}:
            continue
        if c.source_type == "page-list-entry" and c.source_domain in {"www.shixiseng.com", "wap.shixiseng.com"}:
            continue
        if c.source_domain in blocked_domains:
            continue
        if any(bad in c.url.lower() for bad in [".png", ".jpg", ".jpeg", ".gif", "lietou-static"]):
            continue
        if not ("实习" in text or "intern" in text):
            continue
        if c.score < 20:
            continue
        shortlist.append(c)
    return shortlist[:100]


def write_csv(path: Path, rows: list[Candidate]) -> None:
    fields = list(rows[0].as_row().keys()) if rows else []
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        for row in rows:
            writer.writerow(row.as_row())


def write_xlsx(path: Path, rows: list[Candidate]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "岗位短名单"
    headers = list(rows[0].as_row().keys())
    ws.append(headers)
    for c in rows:
        ws.append([c.as_row()[h] for h in headers])

    fill = PatternFill("solid", fgColor="111111")
    for cell in ws[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    widths = {
        "A": 10,
        "B": 18,
        "C": 34,
        "D": 24,
        "E": 18,
        "F": 15,
        "G": 20,
        "H": 44,
        "I": 36,
        "J": 18,
        "K": 16,
        "L": 48,
        "M": 34,
        "N": 55,
        "O": 14,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        row[0].font = Font(bold=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(path)


def write_markdown(path: Path, rows: list[Candidate]) -> None:
    lines = [
        "# 陈毅哲实习岗位短名单 100 条",
        "",
        f"- 采集日期：{TODAY}",
        "- 适配画像：西南财经大学金融学本科 2028 届；GPA 3.7/4.0 前10%；CET-6；7月初可到岗；暑期每周4-5天；项目侧重 AI工作流、公开资料研究、商业/品牌分析。",
        "- 筛选原则：优先 AI运营/AI产品运营/AI Agent工作流、商业/数据分析、品牌/用户/内容运营、金融投研/行研助理；优先远程、成都、有薪资、暑期可投、本科友好。",
        "- 注意：招聘信息变化很快，投递前仍需点开来源确认开放状态、面向届别和出勤要求。",
        "",
        "| # | 分数 | 方向 | 岗位 | 公司/来源 | 地点/远程 | 薪资 | 适配理由 | 风险/备注 | 链接 |",
        "| --- | ---: | --- | --- | --- | --- | --- | --- | --- | --- |",
    ]
    for i, c in enumerate(rows, 1):
        r = c.as_row()
        def esc(x: object) -> str:
            return str(x).replace("|", "/").replace("\n", " ").strip()
        lines.append(
            f"| {i} | {r['rank_score']} | {esc(r['category'])} | {esc(r['title'])} | "
            f"{esc(r['company_or_source'])} | {esc(r['location_remote'])} | {esc(r['salary'])} | "
            f"{esc(r['fit_reason'])} | {esc(r['risk_or_note'])} | [来源]({r['source_url']}) |"
        )
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_summary(rows: list[Candidate], raw: list[Candidate]) -> None:
    by_cat: defaultdict[str, int] = defaultdict(int)
    by_domain: defaultdict[str, int] = defaultdict(int)
    for c in rows:
        by_domain[c.source_domain] += 1
        for cat in c.categories:
            by_cat[cat] += 1
    summary = {
        "collected_at": TODAY,
        "raw_candidates": len(raw),
        "shortlist_count": len(rows),
        "category_counts": dict(sorted(by_cat.items(), key=lambda x: x[1], reverse=True)),
        "source_domain_counts": dict(sorted(by_domain.items(), key=lambda x: x[1], reverse=True)[:30]),
        "outputs": {
            "csv": str(TOP_CSV),
            "xlsx": str(TOP_XLSX),
            "markdown": str(TOP_MD),
            "raw_csv": str(RAW_CSV),
        },
    }
    (OUT_DIR / "internship_shortlist_summary.json").write_text(
        json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8"
    )


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    raw = load_candidates()
    ranked = dedupe(raw)
    shortlist = filter_shortlist(ranked)
    if len(shortlist) < 100:
        print(f"WARNING: only {len(shortlist)} qualified candidates after filtering")
    write_csv(RAW_CSV, ranked)
    write_csv(TOP_CSV, shortlist)
    write_xlsx(TOP_XLSX, shortlist)
    write_markdown(TOP_MD, shortlist)
    write_summary(shortlist, ranked)
    print(f"raw={len(raw)} deduped={len(ranked)} shortlist={len(shortlist)}")
    print(TOP_XLSX)
    print(TOP_MD)


if __name__ == "__main__":
    main()
